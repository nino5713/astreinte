"""
SOCOM - Gestion des astreintes
Application standalone : planning d'astreinte, demandes de dépannage,
statut temps réel des techniciens et compteur horaire (10h/jour, 48h/semaine).

Stack : Flask + SQLite. Conçu pour un déploiement gunicorn + nginx (VPS).
"""

import os
import io
import calendar
import sqlite3
from datetime import datetime, timedelta, time as dtime
from functools import wraps
from zoneinfo import ZoneInfo

from flask import (
    Flask, g, request, session, redirect, url_for,
    render_template, jsonify, abort, send_from_directory, Response, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "astreinte.db")
TZ = ZoneInfo("Europe/Luxembourg")

# Plafonds légaux (modifiables). Le seuil "alerte" déclenche l'avertissement.
MAX_H_JOUR = 10.0
MAX_H_SEMAINE = 48.0
ALERTE_H_JOUR = 8.0        # ambre au-delà
ALERTE_H_SEMAINE = 40.0    # ambre au-delà

STATUTS_TECH = ["disponible", "en_depannage", "repos", "indisponible"]
STATUTS_DEP = ["nouveau", "assigne", "en_cours", "termine", "annule"]
PRIORITES = ["normale", "urgente", "critique"]
ROLES = ["technicien", "dispatcher", "admin"]
# Palette de couleurs distinctes attribuées aux techniciens.
TECH_COULEURS = [
    "#2563EB", "#16A34A", "#EA580C", "#DC2626", "#7C3AED", "#0891B2",
    "#CA8A04", "#DB2777", "#059669", "#4F46E5", "#B45309", "#0D9488",
    "#9333EA", "#65A30D", "#E11D48", "#0369A1", "#C026D3", "#15803D",
]

app = Flask(__name__)
app.secret_key = os.environ.get("ASTREINTE_SECRET", "change-me-en-production")


def _asset_version():
    """Empreinte courte du JS/CSS : change à chaque modification de contenu,
    ce qui casse proprement le cache navigateur/PWA après un déploiement."""
    import hashlib
    h = hashlib.md5()
    for fn in ("static/app.js", "static/style.css"):
        try:
            with open(os.path.join(BASE_DIR, fn), "rb") as f:
                h.update(f.read())
        except OSError:
            pass
    return h.hexdigest()[:8]


ASSET_VERSION = _asset_version()


@app.context_processor
def _injecter_version():
    return {"v": ASSET_VERSION}


# --------------------------------------------------------------------------
# Base de données
# --------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS techniciens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            telephone TEXT,
            role TEXT NOT NULL DEFAULT 'technicien',   -- 'technicien' | 'dispatcher'
            pin_hash TEXT NOT NULL,
            statut TEXT NOT NULL DEFAULT 'disponible',
            couleur TEXT NOT NULL DEFAULT '#64748B',
            actif INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS depannages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client TEXT NOT NULL,
            lieu TEXT,
            description TEXT,
            priorite TEXT NOT NULL DEFAULT 'normale',
            statut TEXT NOT NULL DEFAULT 'nouveau',
            technicien_id INTEGER REFERENCES techniciens(id),
            date_creation TEXT NOT NULL,
            date_debut TEXT,      -- clic "Démarrer l'intervention"
            date_fin TEXT,        -- clic "Terminer"
            cree_par INTEGER REFERENCES techniciens(id)
        );

        CREATE TABLE IF NOT EXISTS astreintes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            technicien_id INTEGER NOT NULL REFERENCES techniciens(id),
            date_debut TEXT NOT NULL,   -- date ISO (jour)
            date_fin TEXT NOT NULL,     -- date ISO (jour), incluse
            libelle TEXT
        );

        CREATE TABLE IF NOT EXISTS equipes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            couleur TEXT NOT NULL DEFAULT '#1E3A8A',
            deux_colonnes INTEGER NOT NULL DEFAULT 0,   -- 1 = colonne Back-up
            heures_jour REAL NOT NULL DEFAULT 0,        -- heures créditées Lun-Ven aux membres
            actif INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS equipe_membres (
            equipe_id INTEGER NOT NULL REFERENCES equipes(id) ON DELETE CASCADE,
            technicien_id INTEGER NOT NULL REFERENCES techniciens(id) ON DELETE CASCADE,
            PRIMARY KEY (equipe_id, technicien_id)
        );

        -- Planning : un technicien d'astreinte par équipe / jour / poste.
        CREATE TABLE IF NOT EXISTS gardes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipe_id INTEGER NOT NULL REFERENCES equipes(id) ON DELETE CASCADE,
            technicien_id INTEGER NOT NULL REFERENCES techniciens(id) ON DELETE CASCADE,
            jour TEXT NOT NULL,                       -- YYYY-MM-DD
            slot TEXT NOT NULL DEFAULT 'titulaire',   -- 'titulaire' | 'backup'
            UNIQUE(equipe_id, jour, slot)
        );
        """
    )
    _migration_astreintes_equipes(db)
    _migration_equipe_double(db)
    _migration_equipe_heures(db)
    _migration_tech_couleur(db)
    # Jeu de données initial (uniquement si vide)
    n = db.execute("SELECT COUNT(*) AS c FROM techniciens").fetchone()["c"]
    if n == 0:
        seed = [
            ("Dispatch", "", "dispatcher", "0000"),
            ("Marco Ferreira", "+352 621 000 001", "technicien", "1111"),
            ("Luc Weber", "+352 621 000 002", "technicien", "2222"),
            ("Ana Silva", "+352 621 000 003", "technicien", "3333"),
            ("Tom Klein", "+352 621 000 004", "technicien", "4444"),
        ]
        for nom, tel, role, pin in seed:
            db.execute(
                "INSERT INTO techniciens (nom, telephone, role, pin_hash) VALUES (?,?,?,?)",
                (nom, tel, role, generate_password_hash(pin)),
            )
        db.commit()
    db.close()


# --------------------------------------------------------------------------
# Utilitaires temps / compteur horaire
# --------------------------------------------------------------------------
def _migration_astreintes_equipes(db):
    """Ajoute equipe_id à astreintes et rend technicien_id nullable.
    Idempotent : ne s'exécute que si la colonne equipe_id est absente."""
    cols = [r[1] for r in db.execute("PRAGMA table_info(astreintes)").fetchall()]
    if "equipe_id" in cols:
        return
    db.executescript(
        """
        CREATE TABLE astreintes_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipe_id INTEGER REFERENCES equipes(id),
            technicien_id INTEGER REFERENCES techniciens(id),
            date_debut TEXT NOT NULL,
            date_fin TEXT NOT NULL,
            libelle TEXT
        );
        INSERT INTO astreintes_new (id, technicien_id, date_debut, date_fin, libelle)
            SELECT id, technicien_id, date_debut, date_fin, libelle FROM astreintes;
        DROP TABLE astreintes;
        ALTER TABLE astreintes_new RENAME TO astreintes;
        """
    )
    db.commit()


def _migration_equipe_double(db):
    """Ajoute la colonne deux_colonnes à equipes si absente (bases existantes)."""
    cols = [r[1] for r in db.execute("PRAGMA table_info(equipes)").fetchall()]
    if "deux_colonnes" not in cols:
        db.execute("ALTER TABLE equipes ADD COLUMN deux_colonnes INTEGER NOT NULL DEFAULT 0")
        db.commit()


def _migration_tech_couleur(db):
    """Ajoute la colonne couleur aux techniciens si absente, et attribue une
    couleur distincte de la palette aux techniciens existants."""
    cols = [r[1] for r in db.execute("PRAGMA table_info(techniciens)").fetchall()]
    if "couleur" in cols:
        return
    db.execute("ALTER TABLE techniciens ADD COLUMN couleur TEXT NOT NULL DEFAULT '#64748B'")
    techs = db.execute("SELECT id FROM techniciens WHERE role = 'technicien' ORDER BY id").fetchall()
    for i, t in enumerate(techs):
        db.execute("UPDATE techniciens SET couleur = ? WHERE id = ?",
                   (TECH_COULEURS[i % len(TECH_COULEURS)], t["id"]))
    db.commit()


def _migration_equipe_heures(db):
    """Ajoute la colonne heures_jour à equipes si absente."""
    cols = [r[1] for r in db.execute("PRAGMA table_info(equipes)").fetchall()]
    if "heures_jour" not in cols:
        db.execute("ALTER TABLE equipes ADD COLUMN heures_jour REAL NOT NULL DEFAULT 0")
        db.commit()


def now_tz():
    return datetime.now(TZ)


def iso(dt):
    return dt.isoformat()


def parse(s):
    if not s:
        return None
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TZ)
    return dt


def _overlap_heures(deb, fin, borne_debut, borne_fin):
    """Durée (heures) de l'intersection [deb,fin] ∩ [borne_debut,borne_fin]."""
    start = max(deb, borne_debut)
    end = min(fin, borne_fin)
    if end <= start:
        return 0.0
    return (end - start).total_seconds() / 3600.0


def bornes_jour(ref_date):
    debut = datetime.combine(ref_date, dtime.min, tzinfo=TZ)
    return debut, debut + timedelta(days=1)


def bornes_semaine(ref_date):
    # Semaine ISO : lundi -> dimanche
    lundi = ref_date - timedelta(days=ref_date.isoweekday() - 1)
    debut = datetime.combine(lundi, dtime.min, tzinfo=TZ)
    return debut, debut + timedelta(days=7)


def heures_technicien(db, tech_id, ref=None):
    """Retourne les heures travaillées sur le jour et la semaine en cours.

    Calcul basé sur les interventions (date_debut -> date_fin). Une
    intervention en cours compte jusqu'à maintenant. Les interventions à
    cheval sur minuit sont découpées correctement (clipping)."""
    if ref is None:
        ref = now_tz()
    ref_date = ref.date()
    jd, jf = bornes_jour(ref_date)
    sd, sf = bornes_semaine(ref_date)

    rows = db.execute(
        """SELECT date_debut, date_fin FROM depannages
           WHERE technicien_id = ? AND date_debut IS NOT NULL
             AND statut IN ('en_cours', 'termine')""",
        (tech_id,),
    ).fetchall()

    h_jour = h_sem = 0.0
    for r in rows:
        deb = parse(r["date_debut"])
        fin = parse(r["date_fin"]) if r["date_fin"] else now_tz()
        h_jour += _overlap_heures(deb, fin, jd, jf)
        h_sem += _overlap_heures(deb, fin, sd, sf)

    # Heures de base (travail régulier) créditées aux membres d'équipe, Lun-Ven.
    taux = db.execute(
        """SELECT COALESCE(SUM(e.heures_jour), 0) AS t
           FROM equipe_membres m JOIN equipes e ON m.equipe_id = e.id
           WHERE m.technicien_id = ? AND e.actif = 1""",
        (tech_id,),
    ).fetchone()["t"]
    if taux:
        iso = ref_date.isoweekday()          # 1 = lundi ... 7 = dimanche
        if iso <= 5:
            h_jour += taux                    # jour ouvré : crédite le jour courant
        h_sem += taux * min(iso, 5)           # semaine : jours ouvrés écoulés (Lun -> aujourd'hui)

    return {"jour": round(h_jour, 2), "semaine": round(h_sem, 2)}


def etat_horaire(h_jour, h_sem):
    """Niveau d'alerte global : ok / attention / depasse."""
    if h_jour >= MAX_H_JOUR or h_sem >= MAX_H_SEMAINE:
        return "depasse"
    if h_jour >= ALERTE_H_JOUR or h_sem >= ALERTE_H_SEMAINE:
        return "attention"
    return "ok"


# --------------------------------------------------------------------------
# Authentification
# --------------------------------------------------------------------------
def courant(db):
    uid = session.get("uid")
    if not uid:
        return None
    return db.execute("SELECT * FROM techniciens WHERE id = ?", (uid,)).fetchone()


def login_requis(f):
    @wraps(f)
    def wrap(*a, **kw):
        if not session.get("uid"):
            if request.path.startswith("/api/"):
                return jsonify({"erreur": "non authentifié"}), 401
            return redirect(url_for("login"))
        return f(*a, **kw)
    return wrap


def dispatcher_requis(f):
    @wraps(f)
    def wrap(*a, **kw):
        db = get_db()
        u = courant(db)
        if not u:
            return redirect(url_for("login"))
        # L'admin est un super-utilisateur : il passe aussi les portes dispatcher.
        if u["role"] not in ("dispatcher", "admin"):
            if request.path.startswith("/api/"):
                return jsonify({"erreur": "réservé au dispatcher"}), 403
            return redirect(url_for("vue_technicien"))
        return f(*a, **kw)
    return wrap


def admin_requis(f):
    @wraps(f)
    def wrap(*a, **kw):
        db = get_db()
        u = courant(db)
        if not u:
            return redirect(url_for("login"))
        if u["role"] != "admin":
            if request.path.startswith("/api/"):
                return jsonify({"erreur": "réservé à l'administrateur"}), 403
            return redirect(url_for("index"))
        return f(*a, **kw)
    return wrap


# --------------------------------------------------------------------------
# Pages
# --------------------------------------------------------------------------
@app.route("/")
def index():
    db = get_db()
    u = courant(db)
    if not u:
        return redirect(url_for("login"))
    if u["role"] == "admin":
        return redirect(url_for("admin"))
    if u["role"] == "dispatcher":
        return redirect(url_for("dashboard"))
    return redirect(url_for("vue_technicien"))


@app.route("/login", methods=["GET", "POST"])
def login():
    db = get_db()
    erreur = None
    if request.method == "POST":
        uid = request.form.get("technicien_id")
        pin = request.form.get("pin", "")
        row = db.execute(
            "SELECT * FROM techniciens WHERE id = ? AND actif = 1", (uid,)
        ).fetchone()
        if row and check_password_hash(row["pin_hash"], pin):
            session["uid"] = row["id"]
            return redirect(url_for("index"))
        erreur = "Code PIN incorrect."
    techs = db.execute(
        "SELECT id, nom, role FROM techniciens WHERE actif = 1 ORDER BY role DESC, nom"
    ).fetchall()
    return render_template("login.html", techniciens=techs, erreur=erreur)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@dispatcher_requis
def dashboard():
    db = get_db()
    return render_template("dashboard.html", utilisateur=courant(db),
                           priorites=PRIORITES)


@app.route("/technicien")
@login_requis
def vue_technicien():
    db = get_db()
    return render_template("technicien.html", utilisateur=courant(db))


@app.route("/planning")
@login_requis
def planning():
    db = get_db()
    return render_template("planning.html", utilisateur=courant(db))


@app.route("/admin")
@admin_requis
def admin():
    db = get_db()
    return render_template("admin.html", utilisateur=courant(db), roles=ROLES)


# --------------------------------------------------------------------------
# API — administration des utilisateurs (réservé admin)
# --------------------------------------------------------------------------
LIB_ROLE = {"technicien": "Technicien", "dispatcher": "Dispatcher", "admin": "Administrateur"}


@app.route("/api/admin/utilisateurs")
@admin_requis
def api_admin_liste():
    db = get_db()
    rows = db.execute(
        "SELECT id, nom, telephone, role, statut, couleur, actif FROM techniciens ORDER BY actif DESC, role DESC, nom"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/admin/utilisateur", methods=["POST"])
@admin_requis
def api_admin_creer():
    db = get_db()
    data = request.get_json(force=True)
    nom = (data.get("nom") or "").strip()
    role = data.get("role", "technicien")
    pin = (data.get("pin") or "").strip()
    tel = (data.get("telephone") or "").strip()
    if not nom:
        return jsonify({"erreur": "Le nom est obligatoire."}), 400
    if role not in ROLES:
        return jsonify({"erreur": "Rôle invalide."}), 400
    if len(pin) < 4:
        return jsonify({"erreur": "Le code PIN doit faire au moins 4 caractères."}), 400
    existe = db.execute("SELECT 1 FROM techniciens WHERE nom = ? AND actif = 1", (nom,)).fetchone()
    if existe:
        return jsonify({"erreur": "Un utilisateur actif porte déjà ce nom."}), 400
    couleur = data.get("couleur")
    if not couleur:
        # Couleur automatique : la prochaine de la palette peu utilisée.
        n = db.execute("SELECT COUNT(*) AS c FROM techniciens WHERE role = 'technicien'").fetchone()["c"]
        couleur = TECH_COULEURS[n % len(TECH_COULEURS)]
    cur = db.execute(
        "INSERT INTO techniciens (nom, telephone, role, pin_hash, statut, couleur, actif) VALUES (?,?,?,?, 'disponible', ?, 1)",
        (nom, tel, role, generate_password_hash(pin), couleur),
    )
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/admin/utilisateur/<int:uid>/couleur", methods=["POST"])
@admin_requis
def api_admin_couleur(uid):
    db = get_db()
    couleur = (request.get_json(force=True).get("couleur") or "").strip()
    if not couleur:
        return jsonify({"erreur": "Couleur requise."}), 400
    if not db.execute("SELECT 1 FROM techniciens WHERE id = ?", (uid,)).fetchone():
        abort(404)
    db.execute("UPDATE techniciens SET couleur = ? WHERE id = ?", (couleur, uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/utilisateur/<int:uid>/pin", methods=["POST"])
@admin_requis
def api_admin_pin(uid):
    db = get_db()
    pin = (request.get_json(force=True).get("pin") or "").strip()
    if len(pin) < 4:
        return jsonify({"erreur": "Le code PIN doit faire au moins 4 caractères."}), 400
    if not db.execute("SELECT 1 FROM techniciens WHERE id = ?", (uid,)).fetchone():
        abort(404)
    db.execute("UPDATE techniciens SET pin_hash = ? WHERE id = ?", (generate_password_hash(pin), uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/utilisateur/<int:uid>/role", methods=["POST"])
@admin_requis
def api_admin_role(uid):
    db = get_db()
    u = courant(db)
    role = request.get_json(force=True).get("role")
    if role not in ROLES:
        return jsonify({"erreur": "Rôle invalide."}), 400
    cible = db.execute("SELECT * FROM techniciens WHERE id = ?", (uid,)).fetchone()
    if not cible:
        abort(404)
    # Ne pas rétrograder le dernier admin actif.
    if cible["role"] == "admin" and role != "admin" and _compte_admins(db, sauf=uid) == 0:
        return jsonify({"erreur": "Impossible : c'est le dernier administrateur actif."}), 400
    db.execute("UPDATE techniciens SET role = ? WHERE id = ?", (role, uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/utilisateur/<int:uid>/actif", methods=["POST"])
@admin_requis
def api_admin_actif(uid):
    db = get_db()
    u = courant(db)
    actif = 1 if request.get_json(force=True).get("actif") else 0
    cible = db.execute("SELECT * FROM techniciens WHERE id = ?", (uid,)).fetchone()
    if not cible:
        abort(404)
    if not actif and uid == u["id"]:
        return jsonify({"erreur": "Vous ne pouvez pas désactiver votre propre compte."}), 400
    if not actif and cible["role"] == "admin" and _compte_admins(db, sauf=uid) == 0:
        return jsonify({"erreur": "Impossible : c'est le dernier administrateur actif."}), 400
    db.execute("UPDATE techniciens SET actif = ? WHERE id = ?", (actif, uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/utilisateur/<int:uid>", methods=["DELETE"])
@admin_requis
def api_admin_supprimer(uid):
    db = get_db()
    u = courant(db)
    cible = db.execute("SELECT * FROM techniciens WHERE id = ?", (uid,)).fetchone()
    if not cible:
        abort(404)
    if uid == u["id"]:
        return jsonify({"erreur": "Vous ne pouvez pas supprimer votre propre compte."}), 400
    if cible["role"] == "admin" and _compte_admins(db, sauf=uid) == 0:
        return jsonify({"erreur": "Impossible : c'est le dernier administrateur actif."}), 400
    # Si l'utilisateur a un historique (dépannages / astreintes), on désactive au lieu de supprimer.
    lien = db.execute(
        "SELECT (SELECT COUNT(*) FROM depannages WHERE technicien_id=?) + (SELECT COUNT(*) FROM astreintes WHERE technicien_id=?) AS n",
        (uid, uid),
    ).fetchone()["n"]
    if lien > 0:
        db.execute("UPDATE techniciens SET actif = 0 WHERE id = ?", (uid,))
        db.commit()
        return jsonify({"ok": True, "desactive": True,
                        "info": "Compte lié à un historique : désactivé plutôt que supprimé."})
    db.execute("DELETE FROM techniciens WHERE id = ?", (uid,))
    db.commit()
    return jsonify({"ok": True, "supprime": True})


def _compte_admins(db, sauf=None):
    if sauf is None:
        return db.execute("SELECT COUNT(*) AS c FROM techniciens WHERE role='admin' AND actif=1").fetchone()["c"]
    return db.execute(
        "SELECT COUNT(*) AS c FROM techniciens WHERE role='admin' AND actif=1 AND id != ?", (sauf,)
    ).fetchone()["c"]


# --------------------------------------------------------------------------
# PWA : service worker (racine), manifeste, page hors-ligne
# --------------------------------------------------------------------------
@app.route("/sw.js")
def service_worker():
    # Servi depuis la racine pour couvrir tout le périmètre de l'app.
    resp = send_from_directory(os.path.join(BASE_DIR, "static"), "sw.js")
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.route("/manifest.webmanifest")
def manifest():
    resp = send_from_directory(os.path.join(BASE_DIR, "static"), "manifest.webmanifest")
    resp.headers["Content-Type"] = "application/manifest+json"
    return resp


@app.route("/offline")
def offline():
    return render_template("offline.html")


# --------------------------------------------------------------------------
# API — état global (polling)
# --------------------------------------------------------------------------
def serialise_depannage(r):
    d = dict(r)
    return d


@app.route("/api/etat")
@login_requis
def api_etat():
    db = get_db()
    u = courant(db)

    techs = db.execute(
        "SELECT * FROM techniciens WHERE actif = 1 ORDER BY role DESC, nom"
    ).fetchall()

    # Gardes du jour : détermine qui est d'astreinte au moment présent.
    aujourdhui = now_tz().date().isoformat()
    gardes = gardes_resolues(db, aujourdhui, aujourdhui)
    oncall = {}  # technicien_id -> liste de postes {equipe_nom, couleur_equipe, slot}
    for g in gardes:
        oncall.setdefault(g["technicien_id"], []).append({
            "equipe_nom": g["equipe_nom"],
            "couleur_equipe": g["equipe_couleur"],
            "slot": g["slot"],
        })

    def carte_tech(t):
        h = heures_technicien(db, t["id"])
        return {
            "id": t["id"], "nom": t["nom"], "telephone": t["telephone"],
            "statut": t["statut"], "couleur": t["couleur"],
            "heures_jour": h["jour"], "heures_semaine": h["semaine"],
            "etat": etat_horaire(h["jour"], h["semaine"]),
            "gardes": oncall.get(t["id"], []),
        }

    # Tableau de bord : uniquement les techniciens d'astreinte aujourd'hui.
    techs_out = [carte_tech(t) for t in techs
                 if t["role"] == "technicien" and t["id"] in oncall]

    # Carte personnelle (pour la vue technicien, qu'il soit d'astreinte ou non).
    moi = None
    if u["role"] == "technicien":
        moi = carte_tech(u)

    deps = db.execute(
        """SELECT d.*, t.nom AS technicien_nom
           FROM depannages d LEFT JOIN techniciens t ON d.technicien_id = t.id
           WHERE d.statut != 'termine' OR d.date_fin >= ?
           ORDER BY
             CASE d.statut WHEN 'en_cours' THEN 0 WHEN 'assigne' THEN 1
                           WHEN 'nouveau' THEN 2 ELSE 3 END,
             CASE d.priorite WHEN 'critique' THEN 0 WHEN 'urgente' THEN 1 ELSE 2 END,
             d.date_creation DESC""",
        (iso(now_tz() - timedelta(hours=12)),),
    ).fetchall()

    return jsonify({
        "maintenant": iso(now_tz()),
        "utilisateur": {"id": u["id"], "nom": u["nom"], "role": u["role"]},
        "techniciens": techs_out,
        "moi": moi,
        "depannages": [serialise_depannage(d) for d in deps],
        "plafonds": {
            "max_jour": MAX_H_JOUR, "max_semaine": MAX_H_SEMAINE,
            "alerte_jour": ALERTE_H_JOUR, "alerte_semaine": ALERTE_H_SEMAINE,
        },
    })


# --------------------------------------------------------------------------
# API — dépannages
# --------------------------------------------------------------------------
@app.route("/api/depannage", methods=["POST"])
@dispatcher_requis
def api_creer_depannage():
    db = get_db()
    u = courant(db)
    data = request.get_json(force=True)
    client = (data.get("client") or "").strip()
    if not client:
        return jsonify({"erreur": "Le client est obligatoire."}), 400
    priorite = data.get("priorite", "normale")
    if priorite not in PRIORITES:
        priorite = "normale"
    tech_id = data.get("technicien_id") or None
    statut = "assigne" if tech_id else "nouveau"
    cur = db.execute(
        """INSERT INTO depannages
           (client, lieu, description, priorite, statut, technicien_id,
            date_creation, cree_par)
           VALUES (?,?,?,?,?,?,?,?)""",
        (client, data.get("lieu", ""), data.get("description", ""),
         priorite, statut, tech_id, iso(now_tz()), u["id"]),
    )
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/depannage/<int:did>/assigner", methods=["POST"])
@dispatcher_requis
def api_assigner(did):
    db = get_db()
    data = request.get_json(force=True)
    tech_id = data.get("technicien_id")
    dep = db.execute("SELECT * FROM depannages WHERE id = ?", (did,)).fetchone()
    if not dep:
        abort(404)
    if dep["statut"] in ("termine", "annule"):
        return jsonify({"erreur": "Dépannage déjà clôturé."}), 400
    db.execute(
        "UPDATE depannages SET technicien_id = ?, statut = 'assigne' WHERE id = ?",
        (tech_id, did),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/depannage/<int:did>/demarrer", methods=["POST"])
@login_requis
def api_demarrer(did):
    db = get_db()
    u = courant(db)
    dep = db.execute("SELECT * FROM depannages WHERE id = ?", (did,)).fetchone()
    if not dep:
        abort(404)
    # Un technicien ne démarre que ses propres interventions ; le dispatcher peut tout.
    if u["role"] != "dispatcher":
        if dep["technicien_id"] and dep["technicien_id"] != u["id"]:
            return jsonify({"erreur": "Ce dépannage est assigné à un autre technicien."}), 403
        tech_id = u["id"]
    else:
        tech_id = dep["technicien_id"]
        if not tech_id:
            return jsonify({"erreur": "Assignez d'abord un technicien."}), 400
    if dep["statut"] in ("termine", "annule"):
        return jsonify({"erreur": "Dépannage déjà clôturé."}), 400
    if dep["statut"] == "en_cours":
        return jsonify({"erreur": "Intervention déjà démarrée."}), 400

    db.execute(
        """UPDATE depannages SET statut = 'en_cours', technicien_id = ?,
           date_debut = ? WHERE id = ?""",
        (tech_id, iso(now_tz()), did),
    )
    db.execute("UPDATE techniciens SET statut = 'en_depannage' WHERE id = ?", (tech_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/depannage/<int:did>/terminer", methods=["POST"])
@login_requis
def api_terminer(did):
    db = get_db()
    u = courant(db)
    dep = db.execute("SELECT * FROM depannages WHERE id = ?", (did,)).fetchone()
    if not dep:
        abort(404)
    if u["role"] != "dispatcher" and dep["technicien_id"] != u["id"]:
        return jsonify({"erreur": "Intervention d'un autre technicien."}), 403
    if dep["statut"] != "en_cours":
        return jsonify({"erreur": "Aucune intervention en cours sur ce dépannage."}), 400

    tech_id = dep["technicien_id"]
    db.execute(
        "UPDATE depannages SET statut = 'termine', date_fin = ? WHERE id = ?",
        (iso(now_tz()), did),
    )
    # Le technicien redevient disponible s'il n'a plus d'intervention en cours.
    reste = db.execute(
        "SELECT COUNT(*) AS c FROM depannages WHERE technicien_id = ? AND statut = 'en_cours'",
        (tech_id,),
    ).fetchone()["c"]
    if reste == 0:
        db.execute("UPDATE techniciens SET statut = 'disponible' WHERE id = ?", (tech_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/depannage/<int:did>/annuler", methods=["POST"])
@dispatcher_requis
def api_annuler(did):
    db = get_db()
    dep = db.execute("SELECT * FROM depannages WHERE id = ?", (did,)).fetchone()
    if not dep:
        abort(404)
    if dep["statut"] == "en_cours" and dep["technicien_id"]:
        reste = db.execute(
            "SELECT COUNT(*) AS c FROM depannages WHERE technicien_id = ? AND statut = 'en_cours' AND id != ?",
            (dep["technicien_id"], did),
        ).fetchone()["c"]
        if reste == 0:
            db.execute("UPDATE techniciens SET statut = 'disponible' WHERE id = ?",
                       (dep["technicien_id"],))
    db.execute("UPDATE depannages SET statut = 'annule' WHERE id = ?", (did,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# API — statut technicien
# --------------------------------------------------------------------------
@app.route("/api/technicien/<int:tid>/statut", methods=["POST"])
@login_requis
def api_statut_tech(tid):
    db = get_db()
    u = courant(db)
    if u["role"] != "dispatcher" and u["id"] != tid:
        return jsonify({"erreur": "Non autorisé."}), 403
    data = request.get_json(force=True)
    statut = data.get("statut")
    if statut not in STATUTS_TECH:
        return jsonify({"erreur": "Statut invalide."}), 400
    # On ne force pas "disponible/repos" si une intervention est en cours.
    en_cours = db.execute(
        "SELECT COUNT(*) AS c FROM depannages WHERE technicien_id = ? AND statut = 'en_cours'",
        (tid,),
    ).fetchone()["c"]
    if en_cours and statut != "en_depannage":
        return jsonify({"erreur": "Terminez l'intervention en cours d'abord."}), 400
    db.execute("UPDATE techniciens SET statut = ? WHERE id = ?", (statut, tid))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# API — planning (gardes) : lecture pour tous, édition admin uniquement
# --------------------------------------------------------------------------
def gardes_resolues(db, debut, fin):
    """Gardes entre [debut, fin] avec équipe et technicien résolus."""
    rows = db.execute(
        """SELECT g.id, g.equipe_id, g.technicien_id, g.jour, g.slot,
                  e.nom AS equipe_nom, e.couleur AS equipe_couleur,
                  t.nom AS technicien_nom, t.couleur AS technicien_couleur
           FROM gardes g
           JOIN equipes e ON g.equipe_id = e.id
           JOIN techniciens t ON g.technicien_id = t.id
           WHERE g.jour >= ? AND g.jour <= ?
           ORDER BY g.jour, e.nom, g.slot""",
        (debut, fin),
    ).fetchall()
    return [dict(r) for r in rows]


@app.route("/api/gardes")
@login_requis
def api_gardes():
    db = get_db()
    debut = request.args.get("debut") or "0000-01-01"
    fin = request.args.get("fin") or "9999-12-31"
    return jsonify(gardes_resolues(db, debut, fin))


@app.route("/api/garde", methods=["POST"])
@admin_requis
def api_definir_garde():
    """Affecte (ou retire) un technicien à une équipe pour un poste, sur un ou
    plusieurs jours consécutifs à partir de 'jour'. technicien_id absent/null
    => on retire l'affectation de ce poste sur la période."""
    db = get_db()
    data = request.get_json(force=True)
    eid = data.get("equipe_id")
    jour = data.get("jour")
    slot = data.get("slot", "titulaire")
    tech_id = data.get("technicien_id")
    if not (eid and jour):
        return jsonify({"erreur": "Équipe et jour requis."}), 400
    if slot not in ("titulaire", "backup"):
        return jsonify({"erreur": "Poste invalide."}), 400
    try:
        base = datetime.strptime(jour, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"erreur": "Date invalide."}), 400
    try:
        nb = max(1, min(int(data.get("jours") or 1), 62))
    except (TypeError, ValueError):
        nb = 1
    if not db.execute("SELECT 1 FROM equipes WHERE id = ?", (eid,)).fetchone():
        abort(404)

    jours = [(base + timedelta(days=k)).isoformat() for k in range(nb)]

    if not tech_id:
        for j in jours:
            db.execute("DELETE FROM gardes WHERE equipe_id = ? AND jour = ? AND slot = ?", (eid, j, slot))
        db.commit()
        return jsonify({"ok": True, "vide": True, "jours": nb})

    # Le technicien doit appartenir à l'équipe.
    membre = db.execute(
        "SELECT 1 FROM equipe_membres WHERE equipe_id = ? AND technicien_id = ?", (eid, tech_id)
    ).fetchone()
    if not membre:
        return jsonify({"erreur": "Ce technicien ne fait pas partie de l'équipe."}), 400

    for j in jours:
        db.execute(
            """INSERT INTO gardes (equipe_id, technicien_id, jour, slot) VALUES (?,?,?,?)
               ON CONFLICT(equipe_id, jour, slot) DO UPDATE SET technicien_id = excluded.technicien_id""",
            (eid, tech_id, j, slot),
        )
    db.commit()
    return jsonify({"ok": True, "jours": nb})


@app.route("/api/equipes")
@login_requis
def api_equipes_light():
    """Équipes actives + membres + option deux colonnes (pour le planning)."""
    db = get_db()
    rows = db.execute(
        "SELECT id, nom, couleur, deux_colonnes, heures_jour FROM equipes WHERE actif = 1 ORDER BY nom"
    ).fetchall()
    out = []
    for e in rows:
        d = dict(e)
        d["membres"] = _membres_equipe(db, e["id"])
        out.append(d)
    return jsonify(out)


# --------------------------------------------------------------------------
# Import / export Excel du planning
# --------------------------------------------------------------------------
def _hex_xl(couleur):
    """#RRGGBB -> FFRRGGBB pour openpyxl."""
    c = (couleur or "#64748B").lstrip("#")
    return "FF" + c.upper()


def _colonnes_planning(db):
    """Liste ordonnée des colonnes (equipe, slot) du planning."""
    equipes = db.execute(
        "SELECT id, nom, couleur, deux_colonnes, heures_jour FROM equipes WHERE actif = 1 ORDER BY nom"
    ).fetchall()
    cols = []
    for e in equipes:
        cols.append((e, "titulaire"))
        if e["deux_colonnes"]:
            cols.append((e, "backup"))
    return equipes, cols


@app.route("/api/planning/export")
@admin_requis
def api_export_planning():
    db = get_db()
    mois = request.args.get("mois") or now_tz().strftime("%Y-%m")
    an, mo = int(mois[:4]), int(mois[5:7])
    nbj = calendar.monthrange(an, mo)[1]
    debut = f"{an:04d}-{mo:02d}-01"
    fin = f"{an:04d}-{mo:02d}-{nbj:02d}"

    equipes, cols = _colonnes_planning(db)
    gardes = gardes_resolues(db, debut, fin)
    idx = {(g["equipe_id"], g["jour"], g["slot"]): (g["technicien_nom"], g["technicien_couleur"]) for g in gardes}

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"
    arial = "Arial"
    bord = Border(*[Side(style="thin", color="FFD0D5DD")] * 4)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    mois_noms = ["", "janvier", "février", "mars", "avril", "mai", "juin",
                 "juillet", "août", "septembre", "octobre", "novembre", "décembre"]
    mois_fr = f"{mois_noms[mo]} {an}"
    ncols = 2 + len(cols)
    ws.cell(1, 1, f"Planning des astreintes — {mois_fr}")
    ws.cell(1, 1).font = Font(name=arial, bold=True, size=14, color="FF1E3A8A")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    # Ligne 2/3 : Date, Jour (fusion verticale)
    for r0, lib in ((1, "Date"), (2, "Jour")):
        pass
    ws.cell(2, 1, "Date"); ws.merge_cells("A2:A3")
    ws.cell(2, 2, "Jour"); ws.merge_cells("B2:B3")

    # En-têtes équipes + sous-postes
    c = 3
    entete_fill = PatternFill("solid", fgColor="FF1E3A8A")
    i = 0
    while i < len(cols):
        e, slot = cols[i]
        col_lettre = get_column_letter(c)
        if e["deux_colonnes"]:
            ws.cell(2, c, e["nom"])
            ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c + 1)
            ws.cell(3, c, "Titulaire")
            ws.cell(3, c + 1, "Back-up")
            for cc in (c, c + 1):
                ws.cell(2, cc).fill = PatternFill("solid", fgColor=_hex_xl(e["couleur"]))
                ws.cell(2, cc).font = Font(name=arial, bold=True, color="FFFFFFFF")
                ws.cell(3, cc).font = Font(name=arial, bold=True, size=9, color="FF475569")
                ws.cell(3, cc).alignment = centre
            c += 2
            i += 2
        else:
            ws.cell(2, c, e["nom"])
            ws.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c)
            ws.cell(2, c).fill = PatternFill("solid", fgColor=_hex_xl(e["couleur"]))
            ws.cell(2, c).font = Font(name=arial, bold=True, color="FFFFFFFF")
            c += 1
            i += 1

    for cell in (ws.cell(2, 1), ws.cell(2, 2)):
        cell.fill = entete_fill
        cell.font = Font(name=arial, bold=True, color="FFFFFFFF")
        cell.alignment = centre

    for cc in range(3, ncols + 1):
        ws.cell(2, cc).alignment = centre

    # Lignes de jours
    jours_fr = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
    for j in range(1, nbj + 1):
        d = datetime(an, mo, j).date()
        r = 3 + j
        ws.cell(r, 1, d.isoformat())
        ws.cell(r, 1).font = Font(name=arial, bold=True)
        ws.cell(r, 2, jours_fr[d.weekday()])
        we = d.weekday() >= 5
        for cc in (1, 2):
            ws.cell(r, cc).alignment = centre
            if we:
                ws.cell(r, cc).fill = PatternFill("solid", fgColor="FFF1F5F9")
        cc = 3
        for (e, slot) in cols:
            val = idx.get((e["id"], d.isoformat(), slot))
            cell = ws.cell(r, cc)
            if val:
                cell.value = val[0]
                cell.fill = PatternFill("solid", fgColor=_hex_xl(val[1]))
                cell.font = Font(name=arial, bold=True, color="FFFFFFFF")
            elif we:
                cell.fill = PatternFill("solid", fgColor="FFF1F5F9")
            cell.alignment = centre
            cc += 1

    # Bordures + largeurs
    for row in ws.iter_rows(min_row=2, max_row=3 + nbj, min_col=1, max_col=ncols):
        for cell in row:
            cell.border = bord
            if not cell.font or cell.font.name != arial:
                cell.font = Font(name=arial)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 11
    for cc in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 16
    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nom_fichier = f"planning_astreinte_{mois}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nom_fichier,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/planning/import", methods=["POST"])
@admin_requis
def api_import_planning():
    db = get_db()
    f = request.files.get("fichier")
    if not f:
        return jsonify({"erreur": "Aucun fichier reçu."}), 400
    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        return jsonify({"erreur": "Fichier Excel illisible."}), 400
    ws = wb.active

    # Repère la ligne d'en-tête (cellule 'Date' en colonne A)
    ligne_equipe = None
    for r in range(1, 8):
        v = ws.cell(r, 1).value
        if v and str(v).strip().lower() == "date":
            ligne_equipe = r
            break
    if not ligne_equipe:
        return jsonify({"erreur": "Format non reconnu : en-tête 'Date' introuvable."}), 400
    ligne_slot = ligne_equipe + 1
    ligne_data = ligne_equipe + 2

    # Cartographie des colonnes -> (nom_equipe, slot)
    equipes = {e["nom"].strip().lower(): e for e in
               db.execute("SELECT id, nom FROM equipes WHERE actif = 1").fetchall()}
    colmap = {}
    equipe_courante = None
    for c in range(3, ws.max_column + 1):
        v = ws.cell(ligne_equipe, c).value
        if v and str(v).strip():
            equipe_courante = str(v).strip()
        slot_lab = ws.cell(ligne_slot, c).value
        slot = "backup" if (slot_lab and "back" in str(slot_lab).lower()) else "titulaire"
        if equipe_courante:
            colmap[c] = (equipe_courante, slot)

    techs = {t["nom"].strip().lower(): t for t in
             db.execute("SELECT id, nom FROM techniciens WHERE actif = 1").fetchall()}

    importes, ignores = 0, []
    for r in range(ligne_data, ws.max_row + 1):
        dval = ws.cell(r, 1).value
        if dval is None:
            continue
        if isinstance(dval, datetime):
            jour = dval.date().isoformat()
        else:
            try:
                jour = datetime.strptime(str(dval).strip()[:10], "%Y-%m-%d").date().isoformat()
            except ValueError:
                continue
        for c, (enom, slot) in colmap.items():
            e = equipes.get(enom.strip().lower())
            if not e:
                continue
            nom = ws.cell(r, c).value
            nom = str(nom).strip() if nom is not None else ""
            if not nom:
                db.execute("DELETE FROM gardes WHERE equipe_id=? AND jour=? AND slot=?", (e["id"], jour, slot))
                continue
            t = techs.get(nom.lower())
            if not t:
                ignores.append(f"{nom} ({jour}) : technicien inconnu")
                continue
            membre = db.execute("SELECT 1 FROM equipe_membres WHERE equipe_id=? AND technicien_id=?",
                                (e["id"], t["id"])).fetchone()
            if not membre:
                ignores.append(f"{nom} ({jour}) : pas membre de {enom}")
                continue
            db.execute(
                """INSERT INTO gardes (equipe_id, technicien_id, jour, slot) VALUES (?,?,?,?)
                   ON CONFLICT(equipe_id, jour, slot) DO UPDATE SET technicien_id = excluded.technicien_id""",
                (e["id"], t["id"], jour, slot),
            )
            importes += 1
    db.commit()
    return jsonify({"ok": True, "importes": importes, "ignores": ignores[:20],
                    "nb_ignores": len(ignores)})


def _heures_valide(v):
    """Normalise les heures/jour d'une équipe : réel entre 0 et 24."""
    try:
        h = float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0
    return round(max(0.0, min(h, 24.0)), 2)


def _membres_equipe(db, eid):
    rows = db.execute(
        """SELECT t.id, t.nom, t.couleur FROM equipe_membres m
           JOIN techniciens t ON m.technicien_id = t.id
           WHERE m.equipe_id = ? ORDER BY t.nom""",
        (eid,),
    ).fetchall()
    return [dict(r) for r in rows]


@app.route("/api/admin/equipes")
@admin_requis
def api_equipes():
    db = get_db()
    equipes = db.execute(
        "SELECT id, nom, couleur, deux_colonnes, heures_jour, actif FROM equipes ORDER BY actif DESC, nom"
    ).fetchall()
    out = []
    for e in equipes:
        d = dict(e)
        d["membres"] = _membres_equipe(db, e["id"])
        out.append(d)
    return jsonify(out)


@app.route("/api/admin/equipe", methods=["POST"])
@admin_requis
def api_creer_equipe():
    db = get_db()
    data = request.get_json(force=True)
    nom = (data.get("nom") or "").strip()
    couleur = data.get("couleur") or "#1E3A8A"
    deux = 1 if data.get("deux_colonnes") else 0
    heures = _heures_valide(data.get("heures_jour"))
    membres = data.get("membres") or []
    if not nom:
        return jsonify({"erreur": "Le nom de l'équipe est obligatoire."}), 400
    cur = db.execute("INSERT INTO equipes (nom, couleur, deux_colonnes, heures_jour) VALUES (?,?,?,?)", (nom, couleur, deux, heures))
    eid = cur.lastrowid
    for tid in membres:
        db.execute("INSERT OR IGNORE INTO equipe_membres (equipe_id, technicien_id) VALUES (?,?)", (eid, tid))
    db.commit()
    return jsonify({"ok": True, "id": eid})


@app.route("/api/admin/equipe/<int:eid>", methods=["POST"])
@admin_requis
def api_modifier_equipe(eid):
    db = get_db()
    if not db.execute("SELECT 1 FROM equipes WHERE id = ?", (eid,)).fetchone():
        abort(404)
    data = request.get_json(force=True)
    nom = (data.get("nom") or "").strip()
    if not nom:
        return jsonify({"erreur": "Le nom de l'équipe est obligatoire."}), 400
    couleur = data.get("couleur") or "#1E3A8A"
    deux = 1 if data.get("deux_colonnes") else 0
    heures = _heures_valide(data.get("heures_jour"))
    db.execute("UPDATE equipes SET nom = ?, couleur = ?, deux_colonnes = ?, heures_jour = ? WHERE id = ?", (nom, couleur, deux, heures, eid))
    if "membres" in data:
        db.execute("DELETE FROM equipe_membres WHERE equipe_id = ?", (eid,))
        for tid in (data.get("membres") or []):
            db.execute("INSERT OR IGNORE INTO equipe_membres (equipe_id, technicien_id) VALUES (?,?)", (eid, tid))
        # Retire du planning les gardes des techniciens qui ne sont plus membres.
        db.execute(
            """DELETE FROM gardes WHERE equipe_id = ? AND technicien_id NOT IN
               (SELECT technicien_id FROM equipe_membres WHERE equipe_id = ?)""",
            (eid, eid),
        )
    # Si on repasse en simple colonne, on retire les gardes back-up.
    if not deux:
        db.execute("DELETE FROM gardes WHERE equipe_id = ? AND slot = 'backup'", (eid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/admin/equipe/<int:eid>", methods=["DELETE"])
@admin_requis
def api_suppr_equipe(eid):
    db = get_db()
    if not db.execute("SELECT 1 FROM equipes WHERE id = ?", (eid,)).fetchone():
        abort(404)
    # Retire aussi les affectations de planning de cette équipe.
    db.execute("DELETE FROM gardes WHERE equipe_id = ?", (eid,))
    db.execute("DELETE FROM astreintes WHERE equipe_id = ?", (eid,))
    db.execute("DELETE FROM equipe_membres WHERE equipe_id = ?", (eid,))
    db.execute("DELETE FROM equipes WHERE id = ?", (eid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/techniciens")
@login_requis
def api_techniciens():
    db = get_db()
    rows = db.execute(
        "SELECT id, nom, telephone, role, statut FROM techniciens WHERE actif = 1 AND role = 'technicien' ORDER BY nom"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# --------------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001)
